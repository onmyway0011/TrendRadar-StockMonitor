import os
import sys
import cv2
import pandas as pd
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

class ImageToExcelConverter:
    def __init__(self):
        # 初始化OCR模型
        self.ocr = PaddleOCR(use_angle_cls=True, lang='ch')

    def process_image(self, image_path):
        # 读取图像
        img = cv2.imread(image_path)
        if img is None:
            return None, "无法读取图像文件"

        # 使用OCR识别
        result = self.ocr.ocr(img, cls=True)
        if not result:
            return None, "未识别到任何内容"

        # 提取文本和坐标
        boxes = []
        texts = []
        for line in result:
            boxes.append(line[0])
            texts.append(line[1][0])

        # 简单的表格结构检测和文本排序
        # 这里使用基于y坐标的简单排序，实际应用可能需要更复杂的表格检测算法
        if len(boxes) == 0:
            return None, "未识别到文本"

        # 按y坐标排序，然后按x坐标排序
        sorted_indices = sorted(range(len(boxes)), key=lambda i: (boxes[i][0][1], boxes[i][0][0]))
        sorted_texts = [texts[i] for i in sorted_indices]
        sorted_boxes = [boxes[i] for i in sorted_indices]

        # 尝试检测行和列
        # 这里使用简单的聚类算法来分组行和列
        # 实际应用可能需要更复杂的表格结构分析
        rows = []
        current_row = [sorted_texts[0]]
        current_y = sorted_boxes[0][0][1]

        for i in range(1, len(sorted_texts)):
            if abs(sorted_boxes[i][0][1] - current_y) < 20:
                current_row.append(sorted_texts[i])
            else:
                rows.append(current_row)
                current_row = [sorted_texts[i]]
                current_y = sorted_boxes[i][0][1]
        rows.append(current_row)

        # 转换为DataFrame
        max_columns = max(len(row) for row in rows)
        for i in range(len(rows)):
            while len(rows[i]) < max_columns:
                rows[i].append("")

        df = pd.DataFrame(rows)
        return df, "识别成功"

    def save_to_excel(self, df, output_path):
        # 创建一个工作簿
        wb = Workbook()
        ws = wb.active

        # 将数据写入工作表
        for r_idx, row in df.iterrows():
            for c_idx, value in enumerate(row):
                ws.cell(row=r_idx+1, column=c_idx+1, value=value)
                # 设置单元格边框
                ws.cell(row=r_idx+1, column=c_idx+1).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # 设置单元格对齐方式
                ws.cell(row=r_idx+1, column=c_idx+1).alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # 保存工作簿
        wb.save(output_path)
        return output_path

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("用法: python image_to_excel.py <图像路径> <输出Excel路径>")
        sys.exit(1)

    image_path = sys.argv[1]
    output_path = sys.argv[2]

    converter = ImageToExcelConverter()
    df, message = converter.process_image(image_path)

    if df is not None:
        excel_path = converter.save_to_excel(df, output_path)
        print(f"成功生成Excel文件: {excel_path}")
    else:
        print(f"错误: {message}")