import os
import sys
import cv2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from image_to_excel import ImageToExcelConverter
import tempfile
import shutil

class VideoToExcelConverter:
    def __init__(self):
        self.image_converter = ImageToExcelConverter()
        self.temp_dir = tempfile.mkdtemp()

    def extract_frames(self, video_path, frame_interval=10):
        # 打开视频文件
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            return None, "无法打开视频文件"

        frame_count = 0
        extracted_frames = []

        # 提取帧
        while True:
            ret, frame = cap.read()
            if not ret:
                break

            # 每隔frame_interval帧提取一帧
            if frame_count % frame_interval == 0:
                frame_path = os.path.join(self.temp_dir, f"frame_{frame_count}.jpg")
                cv2.imwrite(frame_path, frame)
                extracted_frames.append(frame_path)

            frame_count += 1

        cap.release()
        return extracted_frames, "帧提取成功"

    def process_video(self, video_path):
        # 提取帧
        frames, message = self.extract_frames(video_path)
        if not frames:
            return None, message

        # 处理每一帧
        all_results = []
        for frame_path in frames:
            df, msg = self.image_converter.process_image(frame_path)
            if df is not None:
                # 添加帧信息
                df.insert(0, '帧序号', os.path.basename(frame_path).split('_')[1].split('.')[0])
                all_results.append(df)
            else:
                print(f"处理帧 {frame_path} 失败: {msg}")

        if not all_results:
            return None, "没有从视频中识别到有效内容"

        # 合并所有结果
        merged_df = pd.concat(all_results, ignore_index=True)
        return merged_df, "视频处理成功"

    def save_to_excel(self, df, output_path):
        # 创建一个工作簿
        wb = Workbook()
        ws = wb.active

        # 将数据写入工作表
        for r_idx, row in df.iterrows():
            for c_idx, value in enumerate(row):
                ws.cell(row=r_idx+1, column=c_idx+1, value=value)
                # 设置单元格边框
                ws.cell(row=r_idx+1, column=c_idx+1).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # 设置单元格对齐方式
                ws.cell(row=r_idx+1, column=c_idx+1).alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # 保存工作簿
        wb.save(output_path)
        return output_path

    def cleanup(self):
        # 清理临时文件
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("用法: python video_to_excel.py <视频路径> <输出Excel路径>")
        sys.exit(1)

    video_path = sys.argv[1]
    output_path = sys.argv[2]

    converter = VideoToExcelConverter()
    try:
        df, message = converter.process_video(video_path)

        if df is not None:
            excel_path = converter.save_to_excel(df, output_path)
            print(f"成功生成Excel文件: {excel_path}")
        else:
            print(f"错误: {message}")
    finally:
        converter.cleanup()