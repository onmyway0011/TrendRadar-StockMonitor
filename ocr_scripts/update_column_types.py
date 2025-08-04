import os
import sys
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NumberFormatDescriptor

def update_column_types(excel_path, column_types):
    # 加载Excel文件
    try:
        # 使用pandas读取Excel以获取数据
        df = pd.read_excel(excel_path)

        # 加载工作簿和工作表以设置格式
        wb = load_workbook(excel_path)
        ws = wb.active

        # 列类型映射到Excel格式
        format_mapping = {
            'number': '0',
            'decimal': '0.00',
            'currency': '¥#,##0.00',
            'date': 'yyyy-mm-dd',
            'text': '@'
        }

        # 更新列类型
        for column_name, type_name in column_types.items():
            if column_name in df.columns:
                # 获取列索引 (A=1, B=2, ...)
                col_idx = df.columns.get_loc(column_name) + 1
                col_letter = chr(64 + col_idx)  # 将数字转换为字母

                # 设置列格式
                format_code = format_mapping.get(type_name, '@')
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = format_code

        # 保存修改后的文件
        wb.save(excel_path)
        return True, f"成功更新列类型: {excel_path}"
    except Exception as e:
        return False, f"更新列类型时出错: {str(e)}"

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("用法: python update_column_types.py <Excel路径> <列类型JSON>")
        sys.exit(1)

    excel_path = sys.argv[1]
    column_types_json = sys.argv[2]

    try:
        # 解析JSON字符串
        column_types = json.loads(column_types_json)

        # 更新列类型
        success, message = update_column_types(excel_path, column_types)

        if success:
            print(message)
        else:
            print(message)
            sys.exit(1)
    except json.JSONDecodeError:
        print("错误: 无效的JSON格式")
        sys.exit(1)